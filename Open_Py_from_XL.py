import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
# print(wb.get_sheet_names())
# my_sheet = wb.active
# print(my_sheet['A1'].value)
# c = my_sheet['B1']
# print(c.row, c.column, c.value)
# print(c.coordinate, c.value)
# print(my_sheet['C1'].value)
# d = my_sheet.cell(row=1, column=2)

sheet = wb.get_sheet_by_name('Sheet1')
# print(sheet.max_column)
# print(sheet.max_row)
# print(openpyxl.utils.get_column_letter(111))
# print(openpyxl.utils.column_index_from_string ('AA'))

tpl = tuple(sheet['A1':'C3'])
for row_of_cell in tpl:
	for cell in row_of_cell:
		print(cell.value, end=" " )
	print()	

