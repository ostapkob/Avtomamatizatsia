import openpyxl, pprint

wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_active_sheet()
# print(sheet.title)

for i in range(1, sheet.max_column + 1):
    print(sheet.cell(row=1, column=i).value, end=' | ')
print()
for i in range(1, sheet.max_column + 1):
    print(sheet.cell(row=22, column=i).value, end=' | ')
print()
countyData = {}
s = 0
for i in range(2, sheet.max_row + 1):
    state = sheet['B' + str(i)].value
    county = sheet['C' + str(i)].value
    pop = sheet['D' + str(i)].value
    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'track': 0, 'pop': 0})
    countyData[state][county]['track'] += 1
    countyData[state][county]['pop'] += int(pop)
    s += int(pop)
# print(countyData)
with open('dataState.py', 'w') as res:
    res.write('All = ' + pprint.pformat((countyData)))

print(s)
input()