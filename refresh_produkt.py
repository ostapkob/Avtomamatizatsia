import openpyxl
from MyFunctions import print_lists
import os, time

# from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = openpyxl.load_workbook('produceSales.xlsx', data_only=True)
sheet = wb.active
plist = []
max_c = sheet.max_column + 1

price_update = {'Lemon':  1111111111111111111,
                'Celery': 2222222222222222222,
                'Garlic': 3333333333333333333
                }
for strok in range(0, 5):
    plist.append([])
    for i in range(1, max_c):
        plist[strok].append(sheet.cell(row=strok + 1, column=i).value)
print_lists(plist)

u_wb = openpyxl.Workbook()
u_sheet = u_wb.active


undl = openpyxl.styles.Font(name='Calibri',
                            size=14,
                            bold=False,
                            italic=False,
                            vertAlign=None,
                            underline='single',
                            strike=False,
                            color='FF000000')

fil = openpyxl.styles.PatternFill(fill_type='solid',
                                  start_color='000FFFFF',
                                  end_color='FFF00000')

bor = openpyxl.styles.Border(right=openpyxl.styles.Side(border_style='thin',
                                                        color='FF000000'))



for row_num in range(2, 100):
    product = sheet.cell(row=row_num, column=1).value
    for col_num in range(1, max_c):
        u_sheet.cell(row_num, col_num).value = sheet.cell(
            row_num, col_num).value

    if product in price_update:
        u_sheet.cell(row=row_num, column=2).value = price_update[product]
        u_sheet.cell(row=row_num, column=2).font = undl
        u_sheet.cell(row=row_num, column=1).font = openpyxl.styles.Font(
            color='FFFF0000')
        u_sheet.cell(row=row_num, column=3).fill = openpyxl.styles.GradientFill(
            stop=("000000", "FFFFFF"))
        u_sheet.cell(row=row_num, column=4).fill = fil
        u_sheet.cell(row=row_num, column=1).border = bor
        u_sheet.cell(row=row_num, column=5).value = "=SUM(C" + \
            str(row_num) + ": D" + str(row_num) + ")"  # "=SUM(1, 1)"
u_sheet.cell(1, 1).value = 'PRODUCE'
u_sheet.cell(1, 2).value = 'COST PER POUND'
u_sheet.cell(1, 3).value = 'POUNDS SOLD'
u_sheet.cell(1, 4).value = 'TOTAL'

u_sheet.row_dimensions[1].height = 20
u_sheet.column_dimensions['A'].width = 30
u_sheet.merge_cells('B2:B3')
u_sheet.freeze_panes= 'B2'


u_wb.save('udateProduce.xlsx')
os.system('udateProduce.xlsx')
